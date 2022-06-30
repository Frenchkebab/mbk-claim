from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait as wait
import datetime


# 사용자 정의
# from inputFunctions import *
# from mbk import *
# from common import *
# from timer import *

######################### columns.py #########################
column = {
    '1': 'No.',
    '2': 'Commission No.',
    '3': 'VIN No.',
    '4': 'Repair No.',
    '5': 'B/L no.',
    '6': 'Closing Date',
    '7': 'Incident date',
    '8': 'Damage Code01',
    '9': 'Damage Code02',
    '10': 'Damage Code03',
    '11': 'Damage Code04',
    '12': 'Damage Code05',
    '13': 'Damage Code06',
    '14': 'Damage Code07',
    '15': 'Damage Code08',
    '16': 'Damage Code09',
    '17': 'Damage Code10',
    '18': 'Sub Total',
    '19': 'Date of booking',
    '20': 'Reclamation date',
    '21': 'Audit type',
    '22': 'Claim Entry ID-Number',
    '23': 'CID',
    '24': 'Memo'
}
##############################################################


######################### common.py #########################
# from timer import *
import openpyxl
import time
import pyautogui

def altTab():
    pyautogui.keyDown('alt')
    time.sleep(.2)
    pyautogui.press('tab')
    time.sleep(.2)
    pyautogui.keyUp('alt')

def toClaimX(driver):
    driver.get("https://www.claimx.de/claimx")
    driver.implicitly_wait(30)
    

def login(driver, id, password):
    # id 입력
    idForm = driver.find_element(by=By.XPATH, value="/html/body/table/tbody/tr[3]/td[3]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/form/table/tbody/tr[1]/td[2]/input")
    # idForm.send_keys(id)
    # time.sleep(0.5)

    # password 입력
    pwForm = driver.find_element(by=By.XPATH, value="/html/body/table/tbody/tr[3]/td[3]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/form/table/tbody/tr[2]/td[2]/input")
    # pwForm.send_keys(password)

    ## clipboard 사용하여 입력
    clipboard.copy(id)
    idForm.click()
    pyautogui.hotkey('ctrl', 'v')
    time.sleep(0.5)

    clipboard.copy(password)
    pwForm.click()
    pyautogui.hotkey('ctrl', 'v')
    
    print(f"'{id}'")
    print(f"'{password}'")
    input("press Enter: ")
    pwForm.click()
    pwForm.send_keys(Keys.ENTER)
    # 로그인 버튼 클릭
    # driver.find_element(by=By.XPATH, value="/html/body/table/tbody/tr[3]/td[3]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/form/table/tbody/tr[3]/td/input").click()

    # 문 버튼이 나타나면 클릭, 없으면 그냥 패스
    try:
        driver.find_element(by=By.XPATH, value="/html/body/table/tbody/tr[3]/td[3]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/a[1]/img").click()

    except:
        pass

def clickClaim(driver):
    # claim 버튼 클릭
    driver.find_element(by=By.XPATH, value='/html/body/table/tbody/tr[1]/td[3]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table[1]/tbody/tr/td[1]/table/tbody/tr/td[3]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/a').click()

def uploadFileNumCheck(fileList):

    # fileList에서 파악한 총 파일 개수를 더함
    fileCount = 0
    for file in fileList:
        fileCount += len(file)

    return fileCount

def uploadedFileNumCheck(driver):
    print('test: uploadedFileNumbChecker 들어옴')
    wait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, '/html/body/table/tbody/tr[3]/td[1]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr/td/table[5]/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table[17]/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/a'))).click()
    print('test: 화면 리프레시완료')
    uploadedFileNum = wait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//*[@id="mainpart"]/table[6]/tbody/tr[2]/td[2]'))).get_attribute('innerText').strip()
    print('test: ', uploadedFileNum)
    return int(uploadedFileNum)

def memo(file_name, row, msg, archiveError=False):

    # 엑셀 파일 오픈
    wb = openpyxl.load_workbook(f"./upload/{file_name}")

    # 시트 설정
    sheet = wb.worksheets[0]
    sheet.cell(row = 5 + int(row["No."]), column = 24).value = msg

    
    # Memo 작성
    if row == 0:
        sheet.cell(row = 5, column = 24).value = msg
    else:
        # cid값 저장
        sheet.cell(row = 5 + int(row["No."]), column = 24).value = msg


    # archiveError가 있는 경우 Memo 옆에 표시해둠
    if archiveError:
        print("memo 인")
        sheet.cell(row = 5 + int(row["No."]), column = 25).value = "Upload Error"

    # 파일 저장 후 닫기
    wb.save(f"./upload/{file_name}")
    wb.close()

def writeLog(logFile, msg):
    logFile.write(f"{msg}\n")

from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from random import random, randint
import time
import openpyxl

import pyautogui
import clipboard

# from inputFunctions import *
# from common import *
# from timer import *

# cid가 이미 입력되었는지 검사
def checkCID(row):
    # CID가 입력되지 않았으면 
    if str(row["CID"]) == "nan":
        return False

    # 이미 CID가 입력되어 있으면
    else:
        return True


# Vehicle Logistics 입력한 경우 query로 찾아가기
def query(driver, row):
    driver.implicitly_wait(3)
    driver.find_element(by=By.LINK_TEXT, value="query").click() # query 버튼 클릭
    waitLoading()

    vinForm = driver.find_element(by=By.NAME, value="field_akrefitem")
    vinForm.clear()
    time.sleep(0.5)

    cidForm = driver.find_element(by=By.NAME, value="field_aksiditem")
    cidForm.clear()
    time.sleep(1)

    # 만약 경고창이 뜨는 경우
    # try:
    #     driver.find_element(by=By.XPATH, value="/html/body/div[2]/div[1]/button/span[1]").click()
    # except:
    #     pass

    cidForm.send_keys(str(row["CID"]))
    time.sleep(1)
    cidForm.send_keys(Keys.ENTER)

    driver.implicitly_wait(30)
    wait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="mainpart"]/form/table/tbody/tr[2]'))).click()
    # tdTag = driver.find_element(by=By.XPATH, value='//*[@id="mainpart"]/form/table/tbody/tr[2]')
    # tdTag.find_element(by=By.TAG_NAME, value='')
    # uploadBtn.click()

    ### Vehicle Logistics에서 submit한 후의 화면이 나타난다!


def getCid(file_name, driver, row):
    # 텍스트 클릭
    line = driver.find_element(by=By.XPATH, value='//*[@id="show_header_reference"]/table/tbody/tr/td/b').get_attribute('innerText')
    line = line.strip()

    # cid 값 변수 저장
    cid = line[-7:]

    # 엑셀 파일 오픈
    wb = openpyxl.load_workbook(f"./upload/{file_name}")

    # 시트 설정
    sheet = wb.worksheets[0]

    # cid값 저장
    sheet.cell(row = 5 + int(row["No."]), column = 23).value = cid

    # 파일 저장 후 닫기
    wb.save(f"./upload/{file_name}")
    wb.close()

def uploadArchive(driver,logFile, fileList, selectionList):
    div = 1
    for i in range(0, 7):
        for file in fileList[i]:
            writeLog(logFile, file)

            # file selection 클릭
            driver.implicitly_wait(5)
            fileSelection = driver.find_element(by=By.XPATH, value='//*[@id="actions"]/div[1]/span')
            fileSelection.click()
            time.sleep(2)

            # 파일 선택
            pyautogui.write(file)
            time.sleep(2)
            pyautogui.press('enter')
            time.sleep(2)

            # document key 선택
            documentKey = driver.find_element(by=By.XPATH, value=f'//*[@id="previews"]/div[{div}]/div[2]/div[1]/div[2]/div/button')
            div += 1
            action = ActionChains(driver)
            action.move_to_element(documentKey).perform()
            time.sleep(0.5)
            documentKey.click()
            time.sleep(1)

            # document key 입력 후 선택
            clipboard.copy(selectionList[i])
            pyautogui.hotkey('ctrl', 'v')
            time.sleep(0.5)
            pyautogui.press("enter")
            waitLoading()


    driver.find_element(by=By.XPATH, value='//*[@id="actions"]/div[1]/button[1]').click()

    driver.implicitly_wait(5)
    
    # 업로드 화면이 닫혔는지 확인할 때까지 
    while(True):
        time.sleep(2)
        try:
            driver.switch_to.window(driver.window_handles[1])
            continue
        except:
            driver.switch_to.window(driver.window_handles[0])
            waitLoading()
            break
    driver.implicitly_wait(30)
##################################################################################################################################################


########################################### input.py #############################################################################################
# input파일 처리 라이브러리
import pandas as pd
import os
import openpyxl
from email import policy
from email.parser import BytesParser
from glob import glob

def attributeWrite(file_name):
    
    wb = openpyxl.load_workbook(f"./upload/{file_name}")

    sheet = wb.worksheets[0]

    for i in range (1, 25):
        sheet.cell(row = 5, column = i).value = column[f"{i}"]

    wb.save(f"./upload/{file_name}")
    wb.close()

# 엑셀 데이터 입력
def userInput():
    f = open('./settings/setting.txt', 'r', encoding='utf-8')
    id = f.readline().strip()
    password = f.readline().strip()
    f.readline()
    work_type = f.readline().split(':')[1].strip()              # 작업 종류
    file_name = f.readline().split(':')[1].strip()              # 파일 이름
    start_idx = int(f.readline().split(':')[1].strip()) - 1     # 시작 인덱스
    last_idx = int(f.readline().split(':')[1].strip()) - 1      # 마지막 인덱스
    minSecond = int(f.readline().split(':')[1].strip())         # 최소 시간
    maxSecond = int(f.readline().split(':')[1].strip())         # 최대 시간
    f.close()
    
    # 변수 여러개 묶어서 리턴
    user_input = [id, password, work_type, file_name, start_idx, last_idx, minSecond, maxSecond]
    return user_input

def printUserInput(id, password, work_type, file_name, start_idx, last_idx, minSecond, maxSecond):
    print(f"아이디 : {id}")
    print(f"패스워드 : {password}")
    print(f"작업 종류: {work_type}")
    print(f"파일 이름: {file_name}")
    print(f"시작 No.: {start_idx + 1}")
    print(f"마지막 No.: {last_idx + 1}")
    print(f"최소 대기시간(초) : {minSecond}")
    print(f"최대 대기시간(초) : {maxSecond}")

def readXslx(file_name):
    df = pd.read_excel(f'./upload/{file_name}',
                                        skiprows=4,
                                        na_values='',
                                        dtype = {
                                            "No.": str,
                                            "Commission No.": str,
                                            "VIN No.": str,
                                            "Repair No.": str,
                                            "B/L no.": str,
                                            "Closing Date": str,
                                            "Incident date": str,
                                            "Damage Code01": str,
                                            "Damage Code02": str,
                                            "Damage Code03": str,
                                            "Damage Code04": str,
                                            "Damage Code05": str,
                                            "Damage Code05": str,
                                            "Sub Total": str,
                                            "Date of booking": str,
                                            "Reclamation date": str,
                                            "Audit type": str,
                                            "Claim Entry ID-Number": str,
                                            "CID": str
                                        })
    return df

# 데이터프레임의 각 행을 dictionary로 만들고, 각 데이터타입을 정렬
def dfToDictArr(df, start_idx, last_idx):
    dataArr = []
    
    for i in range(start_idx, last_idx + 1):
        rowDict = df.loc[i].to_dict() # 각 행을 딕셔너리로 저장
        rowDict['Closing Date'] = rowDict['Closing Date'][:10]
        rowDict['Incident date'] = rowDict['Incident date'][:10]
        rowDict['Date of booking'] = rowDict['Date of booking'][:10]
        rowDict['Reclamation date'] = rowDict['Reclamation date'][:10]
        dataArr.append(rowDict)

    return dataArr  # 데이터 사전을 원소로 갖는 리스트

def searchFileName(dirName, row):
    currentAbsPath = os.path.dirname(os.path.realpath(__file__))
    dirAbsPath = currentAbsPath + f"\\upload\\{dirName}"
    fileList = os.listdir(dirAbsPath)

    if dirName == "LIST":
        result = [f"{dirAbsPath}\\{fileList[0]}"]
    else:
        result = []
        for file in fileList:
            if file.startswith(row["VIN No."]):
                result.append(f"{dirAbsPath}\\{file}")
        
    return result

def searchEmail(row):
    # 해당 VIN No.와 동일한 EMAIL파일의 경로를 찾는다.
    currentAbsPath = os.path.dirname(os.path.realpath(__file__))
    file_list = list(glob(f"{currentAbsPath}\\upload\\EMAIL\\*.eml"))

    result = []

    for file in file_list:
        with open(file, 'rb') as fp:
            msg = BytesParser(policy=policy.default).parse(fp)
            txt = msg.get_body(preferencelist=('plain')).get_content()
            if txt.find(row["VIN No."]) > -1:
                result.append(file)
    
    return result

##################################################################################################################################################

############################################################## mbk.py #########################################################

def VehicleLogistics(driver, file_name, row):
    # vehicle-logistics로 이동
    driver.implicitly_wait(5)
    driver.find_element(by=By.LINK_TEXT, value="vehicle-logistics").click()
    
    # 페이지 로딩 됐는지 검사
    while True:
        try: 
            driver.find_element(by=By.ID, value="meldfn")
            # 있으면 탈출
            break
        except:
            # 없으면
            time.sleep(5)
            

    # client 선택
    select = Select(driver.find_element(by=By.ID, value="meldfn"))      # Mercedes-Benz Korea Limited
    select.select_by_value("DCD9")
    time.sleep(0.5)

    # product/type of order 선택
    select = Select(driver.find_element(by=By.ID, value="auftrart"))        # day delivery
    select.select_by_value("TAG")
    time.sleep(0.5)

    select = Select(driver.find_element(by=By.NAME, value="field_produktart"))     # transport
    select.select_by_value("TRANS")
    time.sleep(1)
    
    # VIN No. 입력
    first = driver.find_element(by=By.ID, value="sndfzgidwmi")
    first.click()
    time.sleep(0.5)
    first.send_keys(row["VIN No."][0:3])
    time.sleep(0.5)
    second = driver.find_element(by=By.NAME, value="field_sndfzgidvds")
    second.click()
    time.sleep(0.5)
    second.send_keys(row["VIN No."][3:9])
    time.sleep(0.5)
    third = driver.find_element(by=By.NAME, value="field_sndfzgidjahr")
    third.click()
    time.sleep(0.5)
    third.send_keys(row["VIN No."][9:10])
    time.sleep(0.5)
    fourth = driver.find_element(by=By.NAME, value="field_sndfzgidwerk")
    fourth.click()
    time.sleep(0.5)
    fourth.send_keys(row["VIN No."][10:11])
    time.sleep(0.5)
    fifth = driver.find_element(by=By.NAME, value="field_sndfzgidlfd")
    fifth.click()
    time.sleep(0.5)
    fifth.send_keys(row["VIN No."][11:])
    time.sleep(0.5)

    # reference: Commission No.
    commNo = row["Commission No."] 
    if commNo[0] != "0" or len(commNo) != 10:
        commNo = "0" + commNo # Commission No.가 0으로 시작하지 않거나 길이가 10보다 짧으면 0을 붙임

    reference = driver.find_element(by=By.NAME, value="field_auftrref")
    reference.send_keys(row["Commission No."])
    time.sleep(3)

    # further damage 검사
    try:
        driver.find_element(by=By.XPATH, value='//*[@id="mainpart"]/form/table[2]/tbody/tr[7]/td/div/table[1]/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr/td')
        memo(file_name, row, "further damage error")
        return False
    except:
        pass

    # carrier
    carrier = driver.find_element(by=By.NAME, value="field_tuse")
    carrier.send_keys("eukor")
    carrier.send_keys(Keys.ENTER)
    driver.switch_to.window(driver.window_handles[1])
    waitLoading()
    driver.find_element(by=By.XPATH, value="/html/body/table[2]/tbody/tr[2]/td[1]/input").send_keys(Keys.ENTER) # 두 번째 버튼 클릭
    driver.switch_to.window(driver.window_handles[0])

    # reclamation made on
    rYear = row["Reclamation date"][0:4]
    rMonth = row["Reclamation date"][5:7]
    rDay = row["Reclamation date"][-2:]
    driver.find_element(by=By.NAME, value="subfield_reklzeit_day").send_keys(rDay)
    time.sleep(0.5)
    driver.find_element(by=By.NAME, value="subfield_reklzeit_month").send_keys(rMonth)
    time.sleep(0.5)
    driver.find_element(by=By.NAME, value="subfield_reklzeit_year").send_keys(rYear)

    # incident date
    iYear = row["Incident date"][0:4]
    iMonth = row["Incident date"][5:7]
    iDay = row["Incident date"][-2:]
    driver.find_element(by=By.NAME, value="subfield_schzeit_day").send_keys(iDay)
    time.sleep(0.5)
    driver.find_element(by=By.NAME, value="subfield_schzeit_month").send_keys(iMonth)
    time.sleep(0.5)
    driver.find_element(by=By.NAME, value="subfield_schzeit_year").send_keys(iYear)

    # claim type
    Select(driver.find_element(by=By.NAME, value="field_sart")).select_by_value("D01")
    time.sleep(0.5)
    
    # route section/cause
    Select(driver.find_element(by=By.NAME, value="field_sber")).select_by_value("131")
    time.sleep(0.5)
    Select(driver.find_element(by=By.NAME, value="field_surs")).select_by_value("C00")
    time.sleep(0.5)
    
    # claimant's reference
    driver.find_element(by=By.NAME, value="field_ansprref").send_keys(row["Repair No."])
    time.sleep(0.5)

    # policy/type of insurance
    Select(driver.find_element(by=By.NAME, value="field_police")).select_by_value(f"30109636-06154-{iYear}")
    Select(driver.find_element(by=By.NAME, value="field_kzvers")).select_by_value("CL08")

    # estimated/amount claimed
    total = row["Sub Total"]
    driver.find_element(by=By.NAME, value="field_fordmsw").send_keys(total)
    Select(driver.find_element(by=By.NAME, value="field_qmsts")).select_by_value("034")
    time.sleep(0.5)

    # Q-Dome claim?
    driver.find_element(by=By.NAME, value="field_qdome").click()
    time.sleep(0.5)

    # 5-digit-code
    for i in range(1, 11):
        if i < 10:
            dCode = str(row[f"Damage Code0{i}"])
        else:
            dCode = str(row[f"Damage Code{i}"])
        
        # 없으면 반복 중지
        if dCode == "nan":
            break
        else:
            driver.find_element(by=By.NAME, value="field_cteilnr").send_keys(dCode)
            waitLoading()
            driver.find_element(by=By.NAME, value="speichern_ccode").click()
            waitLoading()

    # submit
    driver.find_element(by=By.NAME, value="speichern").click() # submit 버튼 클릭
    driver.implicitly_wait(30)
    waitLoading()

    # return
    return True

def archive(driver, logFile, row, archiveError = False):
    # archive 버튼 클릭
    driver.find_element(by=By.LINK_TEXT, value="archive").click()
    waitLoading()

    # 파일 버튼 클릭
    # driver.find_element(by=By.XPATH, value='//*[@id="mainpart"]/table[4]/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr/td/img').click()
    wait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="mainpart"]/table[5]/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr/td/img'))).click()
    waitLoading()

    # 팝업 창으로
    driver.switch_to.window(driver.window_handles[1])
    
    fRO = searchFileName("RO", row)
    sRO = "claim invoice"

    fBL = searchFileName("BL", row)
    sBL = "B/L"

    fPicture = searchFileName("PICTURE", row)
    sPicture = "Pictures vehicle damage"

    fLiabilityNotice = searchFileName("LIABILITY NOTICE", row)
    sLiabilityNotice = "notice of liability, resp. objection to notice of liability"

    ## damage report는 업로드 안함 (폴더가 비어있음)

    fClaimSummary = searchFileName("CLAIM SUMMARY", row)
    sClaimSummary = "Notification of the claim"

    fEmail = searchEmail(row)
    sEmail = "Incoming correspondence from claimant"

    fList = searchFileName("LIST", row) # 얘만 list 아님
    sList = "Incoming correspondence from claimant"

    fileList = [fRO, fBL, fPicture, fLiabilityNotice, fClaimSummary, fEmail, fList]
    selectionList = [sRO, sBL, sPicture, sLiabilityNotice, sClaimSummary, sEmail, sList]

    # 실제 업로드 로직
    uploadArchive(driver, logFile, fileList, selectionList)

    print("업로드까지는 완료됨")

    # 파악된 파일 개수
    uploadFileNum = uploadFileNumCheck(fileList)

    # 최종 업로드된 파일 개수
    uploadedFileNum = uploadedFileNumCheck(driver)

    # 업로드 상태 확인

    ## 파일의 개수가 안맞는 경우
    if uploadedFileNum != uploadFileNum:
        print("문제1")
        return True

    ## 파일의 개수가 9개 이하인 경우
    if uploadedFileNum < 9:
        print("문제2")
        return True      
    

    # Claim summary = notification of the claim
    # E-mail = Incoming correspondence from claimant
    # Liability notice = Liability notice
    # List = Incoming correspondence from claimant
    # Pictures = pictures 
    # RO = claim invoice

def claim(driver):
    # 좌측 cliam 클릭
    driver.find_element(by=By.XPATH, value="/html/body/table/tbody/tr[3]/td[1]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr/td/table[5]/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table[12]/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/a").click()
    waitLoading()

    # claimant
    claimant = driver.find_element(by=By.NAME, value="field_ansprse")
    claimant.send_keys("Mercedes-Benz Korea")
    time.sleep(0.5)
    claimant.send_keys(Keys.ENTER)

    # 팝업 창
    driver.switch_to.window(driver.window_handles[1])
    waitLoading()

    # + 버튼 클릭
    driver.find_element(by=By.XPATH, value="/html/body/table[2]/tbody/tr[1]/td[1]/input").click()
    driver.implicitly_wait(10)

    # 창 닫혔는지 검사
    while True:
        time.sleep(2)
        try:
            driver.switch_to.window(driver.window_handles[1])
            continue
        except:
            driver.switch_to.window(driver.window_handles[0])
            waitLoading()
            break
    driver.implicitly_wait(30)

    # 창 닫힌 후 submit 버튼 누르기    
    driver.find_element(by=By.NAME, value="speichern").click()
    waitLoading()

    # 완료

def receipts(driver, row):
    # 좌측 receipts 버튼 클릭
    driver.find_element(by=By.XPATH, value='/html/body/table/tbody/tr[3]/td[1]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr/td/table[5]/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table[13]/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/a').click()
    waitLoading()

    # new 버튼 클릭
    driver.find_element(by=By.XPATH, value='//*[@id="mainpart"]/table[1]/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table[1]/tbody/tr/td[2]/a').click()
    waitLoading()

    # type of receipt
    select = Select(driver.find_element(by=By.NAME, value="field_bel"))
    select.select_by_value("RK")
    waitLoading()

    # involved party
    involvedParty = driver.find_element(by=By.NAME, value="field_belansprse")
    involvedParty.send_keys("Mercedes-Benz Korea")
    involvedParty.send_keys(Keys.ENTER)
    
    # 팝업 창 전환
    driver.switch_to.window(driver.window_handles[1])
    driver.implicitly_wait(5)
    waitLoading()

    # + 버튼 클릭
    driver.implicitly_wait(10)
    driver.find_element(by=By.XPATH, value="/html/body/table[2]/tbody/tr[1]/td[1]/input").click()

    # 창 닫혔는지 검사
    while True:
        time.sleep(2)
        try:
            driver.switch_to.window(driver.window_handles[1])
            continue
        except:
            driver.switch_to.window(driver.window_handles[0])
            break
    time.sleep(2)
    
    # receipt number
    driver.find_element(by=By.NAME, value="field_belref").send_keys(row["Repair No."])
    time.sleep(0.5)

    # date of receipt
    receiptYear = row["Closing Date"][0:4]
    receiptMonth = row["Closing Date"][5:7]
    receiptDay = row["Closing Date"][-2:]

    driver.find_element(by=By.NAME, value="subfield_beldat_day").send_keys(receiptDay)
    time.sleep(0.5)
    driver.find_element(by=By.NAME, value="subfield_beldat_month").send_keys(receiptMonth)
    time.sleep(0.5)
    driver.find_element(by=By.NAME, value="subfield_beldat_year").send_keys(receiptYear)
    time.sleep(0.5)

    # tax key
    taxKey = Select(driver.find_element(by=By.NAME, value="field_belstschl"))
    taxKey.select_by_value("100")
    time.sleep(1)

    # amount on receipt nett KRW -> 한 개만 입력하면 나머지는 자동빵
    driver.find_element(by=By.NAME, value="field_betrag_bwhg").send_keys(row["Sub Total"])
    waitLoading()

    # Submit 버튼 클릭
    driver.implicitly_wait(30)
    driver.find_element(by=By.NAME, value="bt_speichern").click()
    waitLoading()

    # >> 버튼 클릭
    driver.find_element(by=By.XPATH, value='//*[@id="mainpart"]/form/table/tbody/tr[3]/td[6]/a[4]/img').click()

    # type of procedure
    typeOfProcedure = Select(driver.find_element(by=By.NAME, value="field_atyp"))
    typeOfProcedure.select_by_value("VR")
    waitLoading()

    # new claim status broker/ins.
    typeOfProcedure = Select(driver.find_element(by=By.NAME, value="field_sst"))
    typeOfProcedure.select_by_value("G")
    waitLoading()

    # delete reserves
    driver.find_element(by=By.NAME, value="field_reskz").click()
    time.sleep(1)

    # date of booking / cost centre
    bookingYear = row["Date of booking"][0:4]
    bookingMonth = row["Date of booking"][5:7]
    bookingDay = row["Date of booking"][-2:]

    bookingDayInput = driver.find_element(by=By.NAME, value="subfield_budat_day")
    bookingDayInput.clear()
    bookingDayInput.send_keys(bookingDay)
    time.sleep(0.5)

    bookingMonthInput = driver.find_element(by=By.NAME, value="subfield_budat_month")
    bookingMonthInput.clear()
    bookingMonthInput.send_keys(bookingMonth)
    time.sleep(0.5)

    bookingYearInput = driver.find_element(by=By.NAME, value="subfield_budat_year")
    bookingYearInput.clear()
    bookingYearInput.send_keys(bookingYear)
    time.sleep(0.5)


    # submit 버튼 클릭
    driver.find_element(by=By.NAME, value="bt_speichern").click()
    waitLoading()

def status(driver):
    # 좌측 status 버튼 클릭
    driver.find_element(by=By.XPATH, value="/html/body/table/tbody/tr[3]/td[1]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr/td/table[5]/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table[18]/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/a").click()
    waitLoading()

    # status
    typeOfProcedure = Select(driver.find_element(by=By.NAME, value="field_sst"))
    typeOfProcedure.select_by_value("B")

    # submit
    driver.find_element(by=By.NAME, value="Abschicken").click()
    waitLoading()

##############################################################


############################## timer.py ################################
import time
from random import random, randint

def sleep_timer_second(min, max):
    range = max - min
    time.sleep(min + random()*range)

def waitLoading():
    time.sleep(2)
##########################################################################

########################## main.py #########################################

# setting.txt 읽기
id, password, work_type, file_name, start_idx, last_idx, minSecond, maxSecond = userInput()

# 로그 작성 파일
now = datetime.datetime.now()
nowDateTime = str(now.strftime('%Y%m%d_%H-%M-%S'))
logFile = open(f"./result/{nowDateTime}.txt", "w", encoding="utf-8")

# 파일이름, 시작NO., 마지막NO. 입력
printUserInput(id, password, work_type, file_name, start_idx, last_idx, minSecond, maxSecond)

# 속성값 덮어쓰기
attributeWrite(file_name)

# 읽은 데이터프레임 받아옴
df = readXslx(file_name)
dataArr = []
for i in range(start_idx, last_idx + 1):
    df.loc[i].to_dict()
    

# 사전 리스트를 저장
dataArr = dfToDictArr(df, start_idx, last_idx)


# 크롬 창 켜기
options = webdriver.ChromeOptions()
options.add_experimental_option("excludeSwitches", ["enable-logging"])
driver = webdriver.Chrome(options=options)

# claimx.com 으로 이동
toClaimX(driver)


# 로그인
input("Login and press enter: ")
# login(driver, id, password)

# Archive 업로드 중 문제가 생기는 경우 True
archiveError = False

# 작업 종류 선택
if work_type == "mbk":
    for row in dataArr:
        try:
            logFile.write(f"========================{row['No.']}번째 줄========================\n")
            print(f"========================{row['No.']}번째 줄========================")
            logFile.write(f"{row}\n\n")
            print(f"{row}\n")

            # 업로드 에러 flag 초기화
            archiveError = False

            #* claim 버튼 클릭
            clickClaim(driver)

            #! CID가 없는 경우 (입력되지 않은 자료)
            if checkCID(row) == False:

                result = VehicleLogistics(driver, file_name, row)
                if result:
                    # 정상적으로 완료되면 CID를 저장
                    getCid(file_name, driver, row)
                    memo(file_name, row, "Vehicle-Logistics done")
                    logFile.write("Vehicle-Logistics 완료\n")
                    print("Vehicle-Logistics 완료")
                else:
                    continue

                # Archive
                archiveError = archive(driver, logFile, row)
                memo(file_name, row, "Archive done", archiveError)
                logFile.write("Archive 완료\n")
                print("Archive 완료")

                # Claim
                claim(driver)
                memo(file_name, row, "Claim done")
                logFile.write("Claim 완료\n")
                print("Claim 완료")

                # Receipts
                receipts(driver, row)
                memo(file_name, row, "Receipts done")
                logFile.write("Receipts 완료\n")
                print("Receipts 완료")

                # Status
                status(driver)
                memo(file_name, row, "finished")
                logFile.write(f"{row['No.']}번 라인 입력 완료\n")
                print(f"{row['No.']}번 라인 입력 완료")
                sleep_timer_second(minSecond, maxSecond)

            #! CID칸이 있는 경우 (입력된 자료)
            else:
                if str(row["Memo"]) == "Vehicle-Logistics done":
                    logFile.write("Archive부터 입력 시작\n")
                    print("Archive부터 입력 시작")
                    query(driver, row)

                    archive(driver,logFile, row)
                    memo(file_name, row, "Archive done")
                    logFile.write("Archive 완료\n")
                    print("Archive 완료")

                    claim(driver)
                    memo(file_name, row, "Claim done")
                    logFile.write("Claim 완료\n")
                    print("Claim 완료")

                    receipts(driver, row)
                    memo(file_name, row, "Receipts done")
                    logFile.write("Receipts 완료\n")
                    print("Receipts 완료")

                    status(driver)
                    memo(file_name, row, "finished")
                    logFile.write(f"{row['No.']}번 라인 입력 완료\n")
                    print(f"{row['No.']}번 라인 입력 완료")

                    logFile.write("============================================================\n\n")
                    print("============================================================\n")


                elif str(row["Memo"]) == "Archive done":
                    logFile.write("Claim 입력부터 시작\n")
                    print("Claim 입력부터 시작")
                    query(driver, row)

                    claim(driver)
                    memo(file_name, row, "Claim done")
                    logFile.write("Claim 완료\n")
                    print("Claim 완료")

                    receipts(driver, row)
                    memo(file_name, row, "Receipts done")
                    logFile.write("Receipts 완료\n")
                    print("Receipts 완료")

                    status(driver)
                    memo(file_name, row, "finished")
                    logFile.write(f"{row['No.']}번 라인 입력 완료\n")
                    print(f"{row['No.']}번 라인 입력 완료")

                elif str(row["Memo"]) == "Claim done":
                    "Receipts 입력부터 시작"
                    query(driver, row)

                    receipts(driver, row)
                    memo(file_name, row, "Receipts done")
                    logFile.write("Receipts 완료\n")
                    print("Receipts 완료")

                    status(driver)
                    memo(file_name, row, "finished")
                    logFile.write(f"{row['No.']}번 라인 입력 완료\n")
                    print(f"{row['No.']}번 라인 입력 완료")

                elif str(row["Memo"]) == "Receipts done":
                    logFile.write("status 입력부터 시작\n")
                    print("status 입력부터 시작")
                    query(driver, row)

                    status(driver)
                    memo(file_name, row, "finished")
                    logFile.write(f"{row['No.']}번 라인 입력 완료\n")
                    print(f"{row['No.']}번 라인 입력 완료")

                elif str(row["Memo"]) == "finished":
                    logFile.write("이미 완료된 작업\n")
                    print("이미 완료된 작업")
                    logFile.write("============================================================\n\n")
                    print("============================================================\n")
                    continue

                else:
                    logFile.write("memo 예외\n")
                    print("memo 예외")
                    logFile.write("============================================================\n\n")
                    print("============================================================\n")
                    continue
            
                sleep_timer_second(minSecond, maxSecond)
            logFile.write("============================================================\n\n")
            print("============================================================\n")
        
        except Exception as e:
            logFile.write(f"{row['No.']} 번째 줄 eror:\n")
            print(f"{row['No.']} 번째 줄 eror:")
            logFile.write(f"{e}\n")
            print(f"{e}")
            logFile.write("============================================================\n\n")
            print("============================================================\n")
            continue

print("프로그램 종료")
logFile.write("프로그램 종료")
logFile.close()