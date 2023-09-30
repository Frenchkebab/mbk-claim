from timer import *
import openpyxl
import time
import pyautogui
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait as wait
import datetime

def altTab():
    pyautogui.keyDown('alt')
    time.sleep(.2)
    pyautogui.press('tab')
    time.sleep(.2)
    pyautogui.keyUp('alt')

def toClaimX(driver):
    driver.get("https://www.claimx.de/claimx")
    driver.implicitly_wait(60 * 20)
    

# def login(driver, id, password):
#     # id 입력
#     driver.find_element(by=By.XPATH, value="/html/body/table/tbody/tr[3]/td[3]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/form/table/tbody/tr[1]/td[2]/input").send_keys(id)
#     time.sleep(0.5)

#     # password 입력
#     driver.find_element(by=By.XPATH, value="/html/body/table/tbody/tr[3]/td[3]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/form/table/tbody/tr[2]/td[2]/input").send_keys(password)
    
#     # 로그인 버튼 클릭
#     driver.find_element(by=By.XPATH, value="/html/body/table/tbody/tr[3]/td[3]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/form/table/tbody/tr[3]/td/input").click()

#     # 문 버튼이 나타나면 클릭, 없으면 그냥 패스
#     try:
#         driver.find_element(by=By.XPATH, value="/html/body/table/tbody/tr[3]/td[3]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/a[1]/img").click()

#     except:
#         pass

def clickClaim(driver):
    # claim 버튼 클릭
    driver.find_element(by=By.XPATH, value='/html/body/table/tbody/tr[1]/td[3]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table[1]/tbody/tr/td[1]/table/tbody/tr/td[3]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/a').click()

def memo(file_name, row, msg):
    # 엑셀 파일 오픈
    wb = openpyxl.load_workbook(f"./upload/{file_name}")

    # 시트 설정
    sheet = wb.worksheets[0]

    if row == 0:
        sheet.cell(row = 5, column = 24).value = msg
    else:
        # cid값 저장
        sheet.cell(row = 5 + int(row["No."]), column = 24).value = msg

    # 파일 저장 후 닫기
    wb.save(f"./upload/{file_name}")
    wb.close()

def writeLog(logFile, msg):
    logFile.write(f"{msg}\n")

# from random import random, randint

#     # policy/type of insurance
#     Select(driver.find_element(by=By.NAME, value="field_police")).select_by_value("30109636-06154-2021")
#     Select(driver.find_element(by=By.NAME, value="field_kzvers")).select_by_value("CL08")

#     # estimated/amount claimed
#     total = row["Sub Total"]
#     driver.find_element(by=By.NAME, value="field_fordmsw").send_keys(total)
#     Select(driver.find_element(by=By.NAME, value="field_qmsts")).select_by_value("034")
#     time.sleep(0.5)

#     # Q-Dome claim?
#     driver.find_element(by=By.NAME, value="field_qdome").click()
#     time.sleep(0.5)

#     # 5-digit-code
#     for i in range(1, 11):
#         if i < 10:
#             dCode = str(row[f"Damage Code0{i}"])
#         else:
#             dCode = str(row[f"Damage Code{i}"])
        
#         # 없으면 반복 중지
#         if dCode == "nan":
#             break
#         else:
#             driver.find_element(by=By.NAME, value="field_cteilnr").send_keys(dCode)
#             waitLoading()
#             driver.find_element(by=By.NAME, value="speichern_ccode").click()
#             waitLoading()

#     # submit
#     driver.find_element(by=By.NAME, value="speichern").click() # submit 버튼 클릭
#     driver.implicitly_wait(60 * 20)
#     waitLoading()

#     # return
#     return True

def archive(driver, logFile, row):
    # archive 버튼 클릭
    driver.find_element_by_link_text("archive").click()
    waitLoading()

    # 파일 버튼 클릭
    driver.find_element(by=By.XPATH, value='//*[@id="mainpart"]/table[5]/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr/td/img').click()
    waitLoading()

    # 팝업 창으로
    driver.switch_to.window(driver.window_handles[1])
    
    fRO = searchFileName("RO", row)
    sRO = "claim invoice"

    fBL = searchFileName("BL", row)
    sBL = "B/L"

    fPicture = searchFileName("PICTURE", row)
    sPicture = "Pictures vehicle damage"

    fLiabilityNotice = searchFileName("LIABILITY NOTICE", row)
    sLiabilityNotice = "notice of liability, resp. objection to notice of liability"

    ## damage report는 업로드 안함 (폴더가 비어있음)

    fClaimSummary = searchFileName("CLAIM SUMMARY", row)
    sClaimSummary = "Notification of the claim"

    fEmail = searchEmail(row)
    sEmail = "Incoming correspondence from claimant"

    fList = searchFileName("LIST", row) # 얘만 list 아님
    sList = "Incoming correspondence from claimant"

    fileList = [fRO, fBL, fPicture, fLiabilityNotice, fClaimSummary, fEmail, fList]
    selectionList = [sRO, sBL, sPicture, sLiabilityNotice, sClaimSummary, sEmail, sList]

    uploadArchive(driver, logFile, fileList, selectionList)

    # Claim summary = notification of the claim
    # E-mail = Incoming correspondence from claimant
    # Liability notice = Liability notice
    # List = Incoming correspondence from claimant
    # Pictures = pictures 
    # RO = claim invoice

def claim(driver):
    # 좌측 cliam 클릭
    driver.find_element(by=By.XPATH, value="/html/body/table/tbody/tr[3]/td[1]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr/td/table[5]/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table[12]/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/a").click()
    waitLoading()

    # claimant
    claimant = driver.find_element(by=By.NAME, value="field_ansprse")
    claimant.send_keys("Mercedes-Benz Korea")
    time.sleep(0.5)
    claimant.send_keys(Keys.ENTER)

    # 팝업 창
    driver.switch_to.window(driver.window_handles[1])
    waitLoading()

    # + 버튼 클릭
    driver.find_element(by=By.XPATH, value="/html/body/table[2]/tbody/tr[1]/td[1]/input").click()
    driver.implicitly_wait(10)

    # 창 닫혔는지 검사
    while True:
        time.sleep(2)
        try:
            driver.switch_to.window(driver.window_handles[1])
            continue
        except:
            driver.switch_to.window(driver.window_handles[0])
            waitLoading()
            break
    driver.implicitly_wait(60 * 20)

    # 창 닫힌 후 submit 버튼 누르기    
    driver.find_element(by=By.NAME, value="speichern").click()
    waitLoading()

    # 완료

def receipts(driver, row):
    # 좌측 receipts 버튼 클릭
    driver.find_element(by=By.XPATH, value='/html/body/table/tbody/tr[3]/td[1]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr/td/table[5]/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table[13]/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/a').click()
    waitLoading()

    # new 버튼 클릭
    driver.find_element(by=By.XPATH, value='//*[@id="mainpart"]/table[1]/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table[1]/tbody/tr/td[2]/a').click()
    waitLoading()

    # type of receipt
    select = Select(driver.find_element(by=By.NAME, value="field_bel"))
    select.select_by_value("RK")
    waitLoading()

    # involved party
    involvedParty = driver.find_element(by=By.NAME, value="field_belansprse")
    involvedParty.send_keys("Mercedes-Benz Korea")
    involvedParty.send_keys(Keys.ENTER)
    
    # 팝업 창 전환
    driver.switch_to.window(driver.window_handles[1])
    driver.implicitly_wait(5)
    waitLoading()

    # + 버튼 클릭
    driver.implicitly_wait(10)
    driver.find_element(by=By.XPATH, value="/html/body/table[2]/tbody/tr[1]/td[1]/input").click()

    # 창 닫혔는지 검사
    while True:
        time.sleep(2)
        try:
            driver.switch_to.window(driver.window_handles[1])
            continue
        except:
            driver.switch_to.window(driver.window_handles[0])
            break
    time.sleep(2)
    
    # receipt number
    driver.find_element(by=By.NAME, value="field_belref").send_keys(row["Repair No."])
    time.sleep(0.5)

    # date of receipt
    receiptYear = row["Closing Date"][0:4]
    receiptMonth = row["Closing Date"][5:7]
    receiptDay = row["Closing Date"][-2:]

    driver.find_element(by=By.NAME, value="subfield_beldat_day").send_keys(receiptDay)
    time.sleep(0.5)
    driver.find_element(by=By.NAME, value="subfield_beldat_month").send_keys(receiptMonth)
    time.sleep(0.5)
    driver.find_element(by=By.NAME, value="subfield_beldat_year").send_keys(receiptYear)
    time.sleep(0.5)

    # tax key
    taxKey = Select(driver.find_element(by=By.NAME, value="field_belstschl"))
    taxKey.select_by_value("100")
    time.sleep(1)

    # amount on receipt nett KRW -> 한 개만 입력하면 나머지는 자동빵
    driver.find_element(by=By.NAME, value="field_betrag_bwhg").send_keys(row["Sub Total"])
    waitLoading()

    # Submit 버튼 클릭
    driver.implicitly_wait(60 * 20)
    driver.find_element(by=By.NAME, value="bt_speichern").click()
    waitLoading()

    # >> 버튼 클릭
    driver.find_element(by=By.XPATH, value='//*[@id="mainpart"]/form/table/tbody/tr[3]/td[6]/a[4]/img').click()

    # type of procedure
    typeOfProcedure = Select(driver.find_element(by=By.NAME, value="field_atyp"))
    typeOfProcedure.select_by_value("VR")
    waitLoading()

    # new claim status broker/ins.
    typeOfProcedure = Select(driver.find_element(by=By.NAME, value="field_sst"))
    typeOfProcedure.select_by_value("G")
    waitLoading()

    # delete reserves
    driver.find_element(by=By.NAME, value="field_reskz").click()
    time.sleep(1)

    # submit 버튼 클릭
    driver.find_element(by=By.NAME, value="bt_speichern").click()
    waitLoading()

def status(driver):
    # 좌측 status 버튼 클릭
    driver.find_element(by=By.XPATH, value="/html/body/table/tbody/tr[3]/td[1]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr/td/table[5]/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table[18]/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/a").click()
    waitLoading()

    # status
    typeOfProcedure = Select(driver.find_element(by=By.NAME, value="field_sst"))
    typeOfProcedure.select_by_value("B")

    # submit
    driver.find_element(by=By.NAME, value="Abschicken").click()
    waitLoading()
# from timer import *
import openpyxl
import time
import pyautogui

def altTab():
    pyautogui.keyDown('alt')
    time.sleep(.2)
    pyautogui.press('tab')
    time.sleep(.2)
    pyautogui.keyUp('alt')

def toClaimX(driver):
    driver.get("https://www.claimx.de/claimx")
    driver.implicitly_wait(30)
    

def login(driver, id, password):
    # id 입력
    idForm = driver.find_element(by=By.XPATH, value="/html/body/table/tbody/tr[3]/td[3]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/form/table/tbody/tr[1]/td[2]/input")
    idForm.send_keys(id)
    time.sleep(0.5)

    # password 입력
    pwForm = driver.find_element(by=By.XPATH, value="/html/body/table/tbody/tr[3]/td[3]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/form/table/tbody/tr[2]/td[2]/input")
    pwForm.send_keys(password)

    ## clipboard 사용하여 입력
    clipboard.copy(id)
    idForm.click()
    pyautogui.hotkey('ctrl', 'v')
    time.sleep(0.5)

    clipboard.copy(password)
    pwForm.click()
    pyautogui.hotkey('ctrl', 'v')
    
    print(f"'{id}'")
    print(f"'{password}'")
    input("press Enter: ")
    pwForm.click()
    pwForm.send_keys(Keys.ENTER)
    # 로그인 버튼 클릭
    driver.find_element(by=By.XPATH, value="/html/body/table/tbody/tr[3]/td[3]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/form/table/tbody/tr[3]/td/input").click()

    # 문 버튼이 나타나면 클릭, 없으면 그냥 패스
    try:
        driver.find_element(by=By.XPATH, value="/html/body/table/tbody/tr[3]/td[3]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/a[1]/img").click()

    except:
        pass

def clickClaim(driver):
    # claim 버튼 클릭
    # driver.find_element(by=By.XPATH, value='/html/body/table/tbody/tr[1]/td[3]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table[1]/tbody/tr/td[1]/table/tbody/tr/td[3]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/a').click()
    driver.find_element(by=By.LINK_TEXT, value="CLAIM").click()

def uploadFileNumCheck(fileList):

    # fileList에서 파악한 총 파일 개수를 더함
    fileCount = 0
    for file in fileList:
        fileCount += len(file)

    return fileCount

def uploadedFileNumCheck(driver):
    wait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, '/html/body/table/tbody/tr[3]/td[1]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr/td/table[5]/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/table[17]/tbody/tr/td/table/tbody/tr/td/table/tbody/tr/td[1]/a'))).click()
    uploadedFileNum = wait(driver, 20).until(EC.presence_of_element_located((By.XPATH, '//*[@id="mainpart"]/table[6]/tbody/tr[2]/td[2]'))).get_attribute('innerText').strip()
    return int(uploadedFileNum)

def memo(file_name, row, msg, archiveError=False):

    # 엑셀 파일 오픈
    wb = openpyxl.load_workbook(f"./upload/{file_name}")

    # 시트 설정
    sheet = wb.worksheets[0]
    sheet.cell(row = 5 + int(row["No."]), column = 24).value = msg

    
    # Memo 작성
    if row == 0:
        sheet.cell(row = 5, column = 24).value = msg
    else:
        # cid값 저장
        sheet.cell(row = 5 + int(row["No."]), column = 24).value = msg


    # archiveError가 있는 경우 Memo 옆에 표시해둠
    if archiveError:
        print("Archive Error Memo")
        sheet.cell(row = 5 + int(row["No."]), column = 25).value = "Upload Error"

    # 파일 저장 후 닫기
    wb.save(f"./upload/{file_name}")
    wb.close()

def writeLog(logFile, msg):
    logFile.write(f"{msg}\n")


from selenium.webdriver import ActionChains
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.by import By
from random import random, randint
import time
import openpyxl

import pyautogui
import clipboard

# from inputFunctions import *
# from common import *
# from timer import *

# cid가 이미 입력되었는지 검사
def checkCID(row):
    # CID가 입력되지 않았으면 
    if str(row["CID"]) == "nan":
        return False

    # 이미 CID가 입력되어 있으면
    else:
        return True


# Vehicle Logistics 입력한 경우 query로 찾아가기
def query(driver, row):
    driver.implicitly_wait(3)
    driver.find_element(by=By.LINK_TEXT, value="QUERY").click() # query 버튼 클릭
    waitLoading()

    vinForm = driver.find_element(by=By.NAME, value="field_akrefitem")
    vinForm.clear()
    time.sleep(0.5)

    cidForm = driver.find_element(by=By.NAME, value="field_aksiditem")
    cidForm.clear()
    time.sleep(1)

    # 만약 경고창이 뜨는 경우
    # try:
    #     driver.find_element(by=By.XPATH, value="/html/body/div[2]/div[1]/button/span[1]").click()
    # except:
    #     pass

    cidForm.send_keys(str(row["CID"]))
    time.sleep(1)
    cidForm.send_keys(Keys.ENTER)

    driver.implicitly_wait(30)
    wait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, '//*[@id="mainpart"]/form/table/tbody/tr[2]'))).click()
    # tdTag = driver.find_element(by=By.XPATH, value='//*[@id="mainpart"]/form/table/tbody/tr[2]')
    # tdTag.find_element(by=By.TAG_NAME, value='')
    # uploadBtn.click()

    ### Vehicle Logistics에서 submit한 후의 화면이 나타난다!


def getCid(file_name, driver, row):
    # 텍스트 클릭
    line = driver.find_element(by=By.XPATH, value='//*[@id="show_header_reference"]/table[1]/tbody/tr/td[1]/div').get_attribute('innerText')
    line = line.strip()

    # cid 값 변수 저장
    cid = line[-7:]

    # 엑셀 파일 오픈
    wb = openpyxl.load_workbook(f"./upload/{file_name}")

    # 시트 설정
    sheet = wb.worksheets[0]

    # cid값 저장
    sheet.cell(row = 5 + int(row["No."]), column = 23).value = cid

    # 파일 저장 후 닫기
    wb.save(f"./upload/{file_name}")
    wb.close()

def uploadArchive(driver,logFile, fileList, selectionList):
    div = 1
    for i in range(0, 7):
        for file in fileList[i]:
            writeLog(logFile, file)

            # file selection 클릭
            driver.implicitly_wait(5)
            fileSelection = driver.find_element(by=By.XPATH, value='//*[@id="actions"]/div[1]/span')
            fileSelection.click()
            time.sleep(2)

            # 파일 선택
            pyautogui.write(file)
            time.sleep(2)
            pyautogui.press('enter')
            time.sleep(2)

            # document key 선택
            documentKey = driver.find_element(by=By.XPATH, value=f'//*[@id="previews"]/div[{div}]/div[2]/div[1]/div[2]/div/button')
            div += 1
            action = ActionChains(driver)
            action.move_to_element(documentKey).perform()
            time.sleep(0.5)
            documentKey.click()
            time.sleep(1)

            # document key 입력 후 선택
            clipboard.copy(selectionList[i])
            pyautogui.hotkey('ctrl', 'v')
            time.sleep(0.5)
            pyautogui.press("enter")
            waitLoading()


    driver.find_element(by=By.XPATH, value='//*[@id="actions"]/div[1]/button[1]').click()

    driver.implicitly_wait(5)
    
    # 업로드 화면이 닫혔는지 확인할 때까지 
    while(True):
        time.sleep(2)
        try:
            driver.switch_to.window(driver.window_handles[1])
            continue
        except:
            driver.switch_to.window(driver.window_handles[0])
            waitLoading()
            break
    driver.implicitly_wait(30)